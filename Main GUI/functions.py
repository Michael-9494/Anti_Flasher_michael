import numpy as np
from scipy import signal
import matplotlib.pyplot as plt
from selenium import webdriver
import pafy
import cv2
import pandas as pd
# from numba import jit
#ffpyplayer for playing audio
from ffpyplayer.player import MediaPlayer
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl import load_workbook



def Save_Data(URL,Time,Freq,Spect,TS_BW,TS_COLOR,TS_FULL, time_frames):
    '''
    Parameters
    ----------
    URL : The website URL
    Time : Time vec for spectogram
    Freq : FREQUENCY vec for spectogram
    Spect :The spectogram Matrix
    TS_BW : Time Stemp for Black % White
    TS_COLOR : Time Stemp for color sensetivity
    TS_FULL : Time Stemp for Full Frame

    This func saves the important data in an excel with a specific loging order

    '''
    print('Saving')
   
    UrlSplit = URL.split("https://www.youtube.com/watch?v=")
    url11 = UrlSplit[1][0:11]    
    temp = url11
    
    #optimizing the URL name to save as a text
    for ch in ['\\','"','*','<','>','#','|','?',':','/','[',']']:
        if ch in temp:
            print(temp)
            temp = temp.replace(ch,"")
      
    dest_filename = 'Data\\Log_Book.xlsx'        
    wb = load_workbook(dest_filename)
    
    ws = wb.create_sheet(title=temp)
    # Rows can also be appended
    ws.append(list(TS_BW))
    ws.append(list(TS_COLOR))
    ws.append(list(TS_FULL))
    ws.append(list(Time))
    ws.append(list(Freq))
    ws.append(list(time_frames))

    for row in range(len(Freq)):
        ws.append(list(Spect[row,:]))
        
    wb.save(filename = dest_filename)
    
def Load_Data(URL):
    '''
    This func load all the Data in 1 sheet according to the sheet name
    Parameters
    ----------
    URL : The URL is the named that is saved by in the Excel file

    Returns
    -------
    Every parameter that had been saved previously will be regenreated 
    '''
    print('Loading')   
    UrlSplit = URL.split("https://www.youtube.com/watch?v=")
    url11 = UrlSplit[1][0:11]
    temp = url11
    
    #optimizing the URL to save as a text
    for ch in ['\\','"','*','<','>','#','|','?',':','/','[',']']:
        if ch in temp:
            temp = temp.replace(ch,"")
    
    dest_filename = 'Data\\Log_Book.xlsx'
    wb = load_workbook(filename=dest_filename)
    wb.active
    
    sh = wb[temp]
    '''
    find TS BW 
    '''
    i = 0
    for row in sh['1']:
        if row.value is None:
            break 
        i += 1
    tsBW = np.zeros((1,i))
    for inx in range(1,i+1):
        char = get_column_letter(inx)
        tsBW[:,inx-1] = sh[char+'1'].value
        
    '''
    find TS COLOR 
    '''
    i = 0
    for row in sh['2']:
        if row.value is None:
            break 
        i += 1
    tsColor = np.zeros((1,i))
    for inx in range(1,i+1):
        char = get_column_letter(inx)
        tsColor[:,inx-1] = sh[char+'2'].value
        
    '''
    find TS Full Frame 
    '''
    i = 0
    for row in sh['3']:
        if row.value is None:
            break 
        i += 1
    tsFull = np.zeros((1,i))
    for inx in range(1,i+1):
        char = get_column_letter(inx)
        tsFull[:,inx-1] = sh[char+'3'].value
    
    '''
    find Time 
    '''
    Spec_col=0
    for row in sh['4']:
        if row.value is None:
            break 
        Spec_col += 1
    time = np.zeros((1,Spec_col))
    for inx in range(1,Spec_col+1):
        char = get_column_letter(inx)
        time[:,inx-1] = sh[char+'4'].value
    '''
    find Freq 
    '''
    Spec_row=0
    for row in sh['5']:
        if row.value is None:
            break 
        Spec_row += 1
    
    freq = np.zeros((1,Spec_row))
    for inx in range(1,Spec_row+1):
        char = get_column_letter(inx)
        freq[:,inx-1] = sh[char+'5'].value  
    
    '''
    find Time Freq 
    '''
    time_freq_inx=0
    for row in sh['6']:
        if row.value is None:
            break 
        time_freq_inx += 1
    
    time_freq = np.zeros((1,time_freq_inx))
    for inx in range(1,time_freq_inx+1):
        char = get_column_letter(inx)
        time_freq[:,inx-1] = sh[char+'6'].value  
        
        
    '''
    find Spect
    '''
    spect = np.zeros((Spec_row,Spec_col))
    for row in range(1,Spec_row+1):
        
        for coll in range(1,Spec_col+1):                
            char = get_column_letter(coll)
            spect[row-1,coll-1] = sh[char+str(row+6)].value
        
    return (tsBW, tsColor, tsFull, time, freq,time_freq, spect)

# @jit( forceobj=True)
def extract_color(frames,R,G,B):
    '''
    extract_color takes Video BGR(blue,green,red) and convert to frames 
    of each color
    
    input:
        frames- 4D array [frameCount, frameWidth, frameHeight,BGR]
        R- 2 the index of Red in BGR colorspace
        G- 1 the index of Green in BGR colorspace
        B= 0 the index of Blue in BGR colorspace
    output:
        frames_r, frames_b, frames_g ,frames_r_minus_b- 
            3D array [frameCount, frameWidth, frameHeight] consists of only 
            one colour, or a difference between red and blue
    examle:
    frames_r, frames_b, frames_g, frames_r_minus_b = functions.extract_color(
        frames)        
        
    '''
    frames_r = frames.copy()[:, :, :, R]
    frames_g = frames.copy()[:, :, :, G]
    frames_b = frames.copy()[:, :, :, B]
    frames_r_minus_b = frames_r - frames_b
    return frames_r, frames_b, frames_g ,frames_r_minus_b


# @jit( forceobj=True)
def cluster_frames_RGB(frames,frameCount,frameHeight,frameWidth,N,dim_for_CLRGB):
    '''
    cluster_frames_RGB creates a N*N clusters of each frame and computes the mean 
    of each cluster throughout the frames.
    
    input:
        frames- 4D array [3,frameCount, frameWidth, frameHeight]
        frameCount- length of video in [samples]
        frameHeight- hight of each frame [pixel]
        frameWidth- width of each frame [pixel]
        N- how many squares each dim
        dim_for_CLRGB- 3 colour space
    output:
        answer- 3D array [frameCount, N, N]
    examle:
        cluster_frameRGB = functions.cluster_frames_RGB(
            frames,frameCount,frameHeight,frameWidth,how_many_squares_each_dim)
        
    '''
    answer = np.empty((dim_for_CLRGB,frameCount, N, N), np.dtype('uint8'))
    for RGB in range(2):    
        for i in range(answer.shape[2]):
            for j in range(answer.shape[3]):
                for k in range(answer.shape[1]):
                    answer[RGB,k,j,i] = np.mean(frames[k,int(j*frameWidth//N):int((j+1)*frameWidth//N),int(i*(frameHeight//N)):int((i+1)*frameHeight//N)])
    return answer

# @jit( forceobj=True)
def dense_cluster_frames(frames,frameCount):
    answer = np.empty((frameCount, 1, 1), np.dtype('uint8'))
    for k in range(answer.shape[0]):
        answer[k,:,:] = np.mean(frames[k,:,:])
        
    vec =  np.empty((frameCount), np.dtype('uint8')) 
    vec = answer[:,0,0]
    return vec


def Open_Chrome():
    print("Im in!")
    # "C:\BGU\Semester 8\Algo\Proj\chromedriver.exe"
    # "C:\Program Files\Google\Chrome\Application\chromedriver.exe"
    Chrome = webdriver.Chrome(executable_path= "C:\Program Files\Google\Chrome\Application\chromedriver.exe")
    return Chrome


def Download_clip(Clip):
    best_quality_video = Clip.getbest()
    ## you can print it to see the quality of the video
    print(best_quality_video)
    ## download it using the download()
    best_quality_video.download()

# @jit( forceobj=True)
def read_vid_to_frames(cap, flag, percent,Percentage ,dim_for_CLRGB,Ms_to_sec ):
    '''
    read_vid_to_frames(cap, flag, percent) is a function that recives the video
    capture, and returns 4D array of the video
    
    input:
        cap-'cv2.VideoCapture' object containing the video
        flag- that is optional to show the video content(in case flag=True 
              while the default is False 
        percent- (number in range of 0-100%) in order to scale the image.
              for example: 50 will make the image size of each frame to be 
              half of the original size 
    output:
        frames- 4D array [frameCount, frameWidth, frameHeight,BGR]
        frameCount- length of video in [samples]
        frameHeight- hight of each frame [pixel]
        frameWidth- width of each frame [pixel]
        fps- sampling frequancy [frames per second]
        Percentage- 100. constant
        dim_for_CLRGB - 3 dimantions
        Ms_to_sec- 10**(-3)
    examle:
        frames, frameCount, frameWidth, frameHeight, fps = functions.read_vid_to_frames(
            cap, False, 50)
    
    '''

    fps = int(cap.get(cv2.CAP_PROP_FPS))
    frameCount = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    frameWidth = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH) * percent / Percentage)
    frameHeight = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT) * percent / Percentage)
    frames = np.empty((frameCount, frameHeight, frameWidth, dim_for_CLRGB ), np.dtype('uint8'))
    time = np.empty((frameCount), np.dtype('float64'))
    # Check if camera opened successfully
    if not cap.isOpened():
        print("Error opening video file")
    fc = 0
    ret = True
    # Read until video is completed
    while fc < frameCount and ret:
        # Capture frame-by-frame
        ret, frame = cap.read()
        # cv2.resize()interpolation=cv2.INTER_AREA)
        frames[fc] = cv2.resize(frame, (frameWidth, frameHeight), interpolation=cv2.INTER_AREA)
        time[fc] = ( Ms_to_sec * cap.get(cv2.CAP_PROP_POS_MSEC))
        fc += 1
        if ret and flag == True:
            # Display the resulting frame
            cv2.imshow('Frame', frames[fc - 1])
            # concede the frames
            # Press Q on keyboard to exit
            if cv2.waitKey(fps) & 0xFF == ord('q'):
                break

    # When everything done, release
    # the video capture object
    cap.release()

    # Closes all the frames
    cv2.destroyAllWindows()
    return (frames, frameCount, frameWidth, frameHeight, fps,time)

def fix_frames(video_path,TS,fps,time_frames,alert_image,time_near_danger):
        '''
        fix_frames(video_path,TS,fps,time_frames,alert_image,time_near_danger) 
        creates a safe version of the video 
       
        input:
          video_path- str with the path to the video
          TS- timestamps found by algorithm
          fps- frames per second [frames/sec]
          time_frames- time vector from the video loading
          alert_image- image to add to dangerous frames
          time_near_danger
        output:
            video of fixed frames
        examle:
            functions.fix_frames(video_path,mainTS_cluster_each_colour_and_find_best,
                  fps,time_frames,alert_image,parameter_list[1][12])
       
        '''

        video=cv2.VideoCapture(video_path)
        player = MediaPlayer(video_path)
        fc = 0
        while True:
            grabbed, frame=video.read()
            audio_frame, val = player.get_frame()
            if not grabbed:
                print("End of video")
                break
            if cv2.waitKey(fps) & 0xFF == ord("q"):
                break
            # in case we are close to the TS from algorithm, change the image
            if (np.abs(time_frames[fc-1]-TS) <= float(time_near_danger)).any():
                new_imRGB =  add_warning_to_im_BGR_ffpyplayer(frame.copy(),alert_image)
                cv2.imshow('Fix-Frame', new_imRGB)
            # second case, no adding needed
            else:
                cv2.imshow('Fix-Frame', frame)
            #audio
            if val != 'eof' and audio_frame is not None:    
                img, t = audio_frame
            fc +=1
            
        video.release()
        cv2.destroyAllWindows()

        return 
    
def spectg(cluster_frame,frameCount,fps,name,flag) :
    N = fps * (1)  # [sample/sec]*[sec]=[sample]
    noverlap = ((50) * N) / 100
    # nf = int((frameCount - N + noverlap) / noverlap);  # the number of segmants
    if flag == 1: 
        plt.figure(figsize=[8, 6])
        powerSpectrum, freqenciesFound, time, imageAxis = plt.specgram(cluster_frame, Fs=fps,noverlap =int(noverlap), NFFT = int(N) )
        plt.xlabel(' Time')
        plt.ylabel('Frequency')
        plt.title(name+" specgram")
        # plt.show()
    else:
        powerSpectrum, freqenciesFound, time,imageAxis = plt.specgram(cluster_frame, Fs=fps,noverlap =int(noverlap), NFFT = int(N) )
        plt.close()
    return(powerSpectrum, freqenciesFound, time, imageAxis)



def Spectrograma_with_print_option(cluster_frame,frameCount,fps,name,flag) :
    '''
    Spectrograma_with_print_option(cluster_frame,frameCount,fps,name,flag) 
    is a function that recives the a vector and computes the spectrograma with 
    emphasis on the time domain, that is because we want to find the time with 
    the optional seizures
    
    input:
        cluster_frame- 1D array that resresent the mean of each cluster 
            accross all frames [frameCount,]
        frameCount- length of video in [samples]
        fps- sampling frequency [frames per second]
        name- string containing the name of the video from youtube.
        flag- when ==1 , the resulting spectrogram will be printed
        
    output:
        powerSpectrum - 2D array containing the spectrogram of the input vector
            (cluster_frame). size [len(freqenciesFound), len(time)]
        freqenciesFound- vector containing the freqencies Found in powerSpectrum
        time- corresponding time vector
        
    examle:
        powerSpectrumRGB, freqenciesFoundRGB, timeRGB= functions.Spectrograma_with_print_option(
            cluster_RGB_vec, frameCount, fps, 'RGB', 1)
    
    '''

    N = fps * (1)  # [sample/sec]*[sec]=[sample]
    noverlap = ((50) * N) / 100
    # nf = int((frameCount - N + noverlap) / noverlap);  # the number of segmants
    if flag == 1: 
        plt.figure(figsize=[8, 6],clear=True)
        freqenciesFound, time, powerSpectrum = signal.spectrogram(
            cluster_frame, fps,nperseg= int(N) , noverlap=int(noverlap),scaling='spectrum', mode ='magnitude' )
        # [V**2]
        # powerSpectrum, freqenciesFound, time, imageAxis = plt.specgram(cluster_frame, Fs=fps,noverlap =int(noverlap), NFFT = int(N) )
        plt.pcolormesh(time, freqenciesFound, powerSpectrum,cmap='coolwarm')
        plt.ylabel('Frequency [Hz]')
        plt.xlabel('Time [sec]')
        plt.title(name+" spectgram")
        # plt.show()
        
    else:
        freqenciesFound, time, powerSpectrum = signal.spectrogram(
            cluster_frame, fps,nperseg= int(N) , noverlap=int(noverlap) ,scaling='spectrum', mode ='magnitude' )
        plt.close()
    return(powerSpectrum, freqenciesFound, time)



def find_time_stamps(powerSpectrum,freqenciesFound,time,fps):
    '''
    find_time_stamps(powerSpectrum,freqenciesFound,time,fps)
    is a function that recives powerSpectrum and creates data frame to remove all
    frequencies Below (the mean power across the minimum power in each time) so
    we will remain with the dangerous 
    
    input:
        powerSpectrum-2D array containing the spectrogram with size
                [len(freqenciesFound), len(time)]
        freqenciesFound- vector containing the freqencies Found in powerSpectrum
        time- corresponding time vector       
        fps- sampling frequency [frames per second]
      
    output:
        time_with_seazure_prob- data Frame containing powerSpectrum above threshold
        TS- time stamps of frames with sizure warning
        
    examle:
        RGB_time_with_seazure_prob, time_stampsRGB_dense_way = functions.find_time_stamps(
            powerSpectrumRGB, freqenciesFoundRGB, timeRGB, fps)
    
    '''
    # convert to data frame
    df = pd.DataFrame(data=powerSpectrum,index = freqenciesFound , columns = time)
    # calculate the minimum power in each time (columns)
    dfmin_mean = np.mean(df.min(axis=0))
    # define threshold
    threshMinPower = int(dfmin_mean *3)
    # threshold the data frame
    freq_above_mat = df[df>threshMinPower]
    threshMaxNan = (fps//2)//3 # how many non NaN we want to keep
    time_with_seazure_prob = freq_above_mat.dropna(thresh= threshMaxNan,axis=1,how='all')
    TS = time_with_seazure_prob.columns.to_numpy()
    return(time_with_seazure_prob,TS)


def cluster_time_stemps(cluster_frameBB,frameCount,fps,dim):
    '''
    cluster_time_stemps(cluster_frameBB,frameCount,fps,dim)
    is a function that recives 3D array [frameCount, dim//2, dim//2] for each 
    vector [:, i, j] computes the spectrogram and find the time stemps. the the
    function goes over the time stamps we found and saving time stemps that apear
    more then 75 % in all the time stemps found
    
    input:
        cluster_frameBB- 3D array [frameCount, dim//2, dim//2] 
        frameCount- length of video in [samples]
        fps- sampling frequency [frames per second]
        dim- how many squares in each frame
        
    output:
        time_with_seazure_prob- data Frame containing powerSpectrum above threshold
        TS- time stamps of frames with sizure warning
        
    examle:
        TS,time_with_seazure_prob = functions.cluster_time_stemps(cluster_frameRGB[1,:,:,:],frameCount,fps,how_many_squares)
    
    '''
    power = []
    df_with_danger_freq = []
    time_stamps_3by3 = []
    time_with_seazure_prob = []
    time_stampss = []
    for i in range(dim//2):
        for j in range(dim//2):
            # compute the spectrogram
            powerSpectrum, freqenciesFound, timee,  = Spectrograma_with_print_option(
                cluster_frameBB[:, i-1, j-1], frameCount, fps, 'RGB', 0)
            power.append(powerSpectrum)
            # timee = timee - 0.5 # shift time vector
            # find time stamps
            time_with_seazure_prob, time_stampss = find_time_stamps(
                powerSpectrum, freqenciesFound, timee, fps)
            df_with_danger_freq.append(time_with_seazure_prob)
            time_stamps_3by3.append(time_stampss)
            

    temp = []
    mainTS = []
    count = 0
    boolvec = []
    for i in range(len(timee)):
        for j in range(len(time_stamps_3by3)):
            if timee[i] in time_stamps_3by3[j]:
                count += 1
                boolvec.append(count)
                temp.append(timee[i])
            else:
                count = 0
                boolvec.append(count) 
        if count >= 0.75*dim:
            mainTS.append(timee[i])  
            
    return(np.array(mainTS),timee)



def cluster_time_stempsRGB(cluster_frame,frameCount,fps,dim,Clip_Name):
    '''
    cluster_time_stemps(cluster_frameBB,frameCount,fps,dim)
    is a function that recives 4D array [RGB(3),frameCount, dim//2, dim//2] for each 
    vector [k,:, i, j] computes the spectrogram and find the time stemps.
    then the function goes over the time stamps we found and saving time
    stemps that apear more then 50 % in all the time stemps found
    
    input:
        cluster_frameBB- 4D array [RGB(3),frameCount, dim//2, dim//2] 
        frameCount- length of video in [samples]
        fps- sampling frequency [frames per second]
        dim- how many squares in each frame
        Clip_Name- string containing the name of the video from youtube
    output:
        countvec- counter
        time_stamps_dim_by_dim- list containing the time stemps found ro each cluster
        np.array(mainTS)- final time stamps
        timee- original time vector of the video
        
    examle:
        countvec,time_stamps_4_by_4,mainTS_RGB_way_1,time = functions.cluster_time_stempsRGB(
            cluster_frameRGB,frameCount,fps,how_many_squares,Clip_Name)
    '''
    power = []
    df_with_danger_freq = []
    time_stamps_dim_by_dim = []
    time_with_seazure_prob = []
    time_stampss = []
    for RGB in range(2):
        for i in range(dim//2):
            for j in range(dim//2):
                powerSpectrum, freqenciesFound, timee = Spectrograma_with_print_option(
                    cluster_frame[RGB,:, i-1, j-1], frameCount, fps, Clip_Name, 0)
                # timee = timee - 0.5 # shift time vector
                power.append(powerSpectrum)
                time_with_seazure_prob, time_stampss = find_time_stamps(
                    powerSpectrum, freqenciesFound, timee, fps)
                df_with_danger_freq.append(time_with_seazure_prob)
                time_stamps_dim_by_dim.append(time_stampss)

    temp = []
    mainTS = []
    count = 0
    countvec = []
    newDim = dim*3
    for i in range(len(timee)):
        for j in range(len(time_stamps_dim_by_dim)):
            if timee[i] in time_stamps_dim_by_dim[j]:
                count += 1
                countvec.append(count)
                temp.append(timee[i])
                if count == newDim:
                    count = 0
            else:
                count = 0
                countvec.append(count) 
        if count >= 0.5*newDim:
            mainTS.append(timee[i])  
            
    return(countvec,time_stamps_dim_by_dim,np.array(mainTS),timee)
  
            
def Take_best_TS(time_stamps_Three_channel,timee):
      
    temp = []
    mainTS = []
    count = 0
    boolvec = []
    for i in range(len(timee)):
        for j in range(len(time_stamps_Three_channel)):
            if timee[i] in time_stamps_Three_channel[j]:
                count += 1
                boolvec.append(count)
                temp.append(timee[i])
            else:
                count = 0
                boolvec.append(count) 
        if count >= 2:
            mainTS.append(timee[i])
            
    return(np.array(mainTS))

def Time_stemp_analysis(TS,time):
    answer = []
    deff = time[1]-time[0]
    for i in range(len(TS)):
        if deff == TS[i]-TS[i-1]:
            print('yes')
            answer.append(TS[i])
        else:
            print('no')
    
    return(answer)

def add_warning_to_im_BGR_ffpyplayer(im,img_ref):
    '''
   add_warning_to_im_BGR_ffpyplayer(im,img_ref)
    is a function that recives 4D array [RGB(3),frameCount, dim//2, dim//2] for each 
    vector [k,:, i, j] computes the spectrogram and find the time stemps.
    then the function goes over the time stamps we found and saving time
    stemps that apear more then 50 % in all the time stemps found
    
    input:
        im- the current frame we want to edit
        img_ref- warning image to print
    output:
        answerBGR- the updated frame
        
    examle:
       new_imBGR =  add_warning_to_im_BGR_ffpyplayer(frame.copy(),alert_image)
    '''
    # resize the warning image
    img_ref_resize =cv2.resize(img_ref, (im.shape[1]//2, im.shape[0]//2), interpolation=cv2.INTER_AREA)
    # preallocate the output
    ref_image_like_im = np.zeros_like(im)
    # broadcast the warning image to answer     
    ref_image_like_im[:img_ref_resize.shape[0],:img_ref_resize.shape[1],:] = img_ref_resize
    # delete space in original image before adding 
    im[:img_ref_resize.shape[0],:img_ref_resize.shape[1],:]=0
    # add new image
    answerBGR =ref_image_like_im +im
    
    return(answerBGR)


    
    
def extract_Hyper_param(file_name):
    '''
    extract_Hyper_param(file_name) is a function for the Hyper parameters of 
    the project. 
    
    input:
        file_name- string with the name of .txt File
    output:
        parameter_list- nested list with 
        [parameter_name[:]- list of parameters names
         ,parameter_val[:]- list of floats containing the values
         ,parameter_explain[:]- list of parameters explanation] 
        
    examle:
       parameter_list = functions.extract_Hyper_param("HyperParam.txt")
    '''
    Hyper_param = open(file_name,"r+")
    params = Hyper_param.readlines()
    parameter_name=[]
    parameter_val =[]
    parameter_explain =[]
    parameter_list = []
    i=0
    for line in params:
        parameter_name.append(params[i].split("=")[0])
        parameter_val.append(float(params[i].split("=")[-1].split("#")[0]))
        parameter_explain.append(params[i].split("=")[-1].split("#")[1])
        i +=1
    parameter_list.append(parameter_name)
    parameter_list.append(parameter_val)
    parameter_list.append(parameter_explain)
    return(parameter_list)
           
    
    
    
    
    
    
    
    
    
    
    
    
    